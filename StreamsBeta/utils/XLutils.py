import openpyxl

class userdata:

    def getrowcount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return (sheet.max_row)

    def getcolumncount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return (sheet.max_column)

    def writestatus(self, path, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(path)

    def getuser(self, path, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rownum, column=columnno).value

